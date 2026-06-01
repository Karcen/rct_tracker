# -*- coding: utf-8 -*-
"""台账导出 Excel 模块（兼容 SPSS 等统计软件）。

导出文件包含 4 个工作表：
- Data        机器友好：英文变量名 + 编码值（M/F、A/B、valid/void...），SPSS 直接读取
- Ledger      人类友好：本地化表头 + 中文/英文标签
- Statistics  统计汇总
- Dictionary  数据字典（编码 -> 含义对照）
"""

import os
from datetime import datetime

from app import config, i18n, db


def export_ledger(path=None):
    """导出全量台账，返回导出文件路径。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except Exception:
        raise RuntimeError("未安装 openpyxl，请先执行：pip install openpyxl")

    rows = db.all_subjects()
    wb = Workbook()

    # ---------- Sheet 1: Data（SPSS 友好，编码值）----------
    ws = wb.active
    ws.title = "Data"
    var_names = ["subject_id", "barcode", "name", "phone", "gender", "group",
                 "register_time", "round1", "round2", "round3", "note"]
    ws.append(var_names)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([
            r["subject_id"], r["barcode"], r["name"], r["phone"], r["gender"],
            r["group_type"], r["register_time"],
            r["round1_status"], r["round2_status"], r["round3_status"], r["note"],
        ])

    # ---------- Sheet 2: Ledger（人类友好，本地化标签）----------
    ws2 = wb.create_sheet("Ledger")
    headers = [i18n.T("col_subject_id"), i18n.T("col_barcode"), i18n.T("col_name"),
               i18n.T("col_phone"), i18n.T("col_gender"), i18n.T("col_group"),
               i18n.T("col_register_time"), i18n.T("col_round1"), i18n.T("col_round2"),
               i18n.T("col_round3"), i18n.T("col_note")]
    ws2.append(headers)
    for c in ws2[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws2.append([
            r["subject_id"], r["barcode"], r["name"], r["phone"],
            i18n.gender_label(r["gender"]), i18n.group_label(r["group_type"]),
            r["register_time"], i18n.status_label(r["round1_status"]),
            i18n.status_label(r["round2_status"]), i18n.status_label(r["round3_status"]),
            r["note"],
        ])

    # ---------- Sheet 3: Statistics ----------
    ws3 = wb.create_sheet("Statistics")
    s = db.get_stats()
    ws3.append([i18n.T("stat_total"), s["total"]])
    ws3.append([i18n.T("stat_group_a"), s["group_a"]])
    ws3.append([i18n.T("stat_group_b"), s["group_b"]])
    ws3.append([])
    head = [i18n.T("stat_round_col"), i18n.T("stat_valid"), i18n.T("stat_void"),
            i18n.T("stat_mismatch"), i18n.T("stat_damaged"), i18n.T("stat_lost"),
            i18n.T("stat_pending")]
    ws3.append(head)
    for c in ws3[5]:
        c.font = Font(bold=True)
    for rnd in (1, 2, 3):
        cc = s["rounds"][rnd]
        ws3.append([i18n.T("round_n", n=rnd), cc["valid"], cc["void"],
                    cc["mismatch"], cc["damaged"], cc["lost"], cc[""]])

    # ---------- Sheet 4: Dictionary（数据字典）----------
    ws4 = wb.create_sheet("Dictionary")
    ws4.append(["field", "code", "meaning_zh / meaning_en"])
    for c in ws4[1]:
        c.font = Font(bold=True)
    dict_rows = [
        ("gender", "M", "男 / Male"), ("gender", "F", "女 / Female"),
        ("gender", "O", "其他 / Other"),
        ("group", "A", "干预组 / Intervention"), ("group", "B", "对照组 / Control"),
        ("round*", "", "未随访 / Not yet"), ("round*", "valid", "正常有效 / Valid"),
        ("round*", "void", "空白作废 / Void"), ("round*", "mismatch", "信息不符 / Mismatch"),
        ("round*", "damaged", "条码损坏 / Damaged"), ("round*", "lost", "失访 / Lost"),
    ]
    for dr in dict_rows:
        ws4.append(list(dr))

    # 适当加宽列宽，便于阅读
    for sheet in (ws, ws2, ws3, ws4):
        for col_cells in sheet.columns:
            width = 12
            for cell in col_cells:
                if cell.value is not None:
                    width = max(width, min(40, len(str(cell.value)) + 2))
            sheet.column_dimensions[col_cells[0].column_letter].width = width
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal="center")

    if not path:
        fn = "rct_ledger_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(config.EXPORT_DIR, fn)
    wb.save(path)
    return path
